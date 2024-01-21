#!/usr/bin/python

from asyncio.log import logger
import sys
import xlrd
import getopt
from collections import defaultdict
import xlsxwriter
import logging as log
from nested_dict import nested_dict
import logging
import sys
import os
from guizero import App, CheckBox, PushButton, Text, TextBox
import matplotlib
import openpyxl
from enum import Enum

# from openpyxl import styles
import subprocess

# constants
OUTPUT_LOG_FILE = "output.txt"

# globals
output_file = ""
input_folder = ""
logger = None
input_file = ""

# dictionary to store the parsed input data
# this is a 3 level nested dictionary such as:
# - Each color per column per sheet from the input needs
#   to go to a new column in the output.
# - The same color from all the different input sheets should
#   go to the same sheet in the output.
# input_dic_parsed[sheet number][background color][column] = []
input_dic_parsed = nested_dict(3, list)

# A map that track the output sheet number for a given color
color_to_sheet_num_map = {}

# A record of all the colors seen, to track the column index
# for any given color, for the output.
# Requirement: Each unique color in an input column should go
# to the same column in the output.
cur_column_per_color = {}

# By default break on whitespaces
break_on_white = True

# Open output file once sorted
open_output_file_once_sorted = True

"""
------------------------------------------------------------
At the moment, there is no single module that supports both
.xls and .xlsx. xlrd doesn't support xlsx and openpyxl doesn't
support .xls. The below section abstracts these module specific
logic into helper routines so that the main logic can remain
the same.
Whether to use xlrd module or openpyxl is determined at runtime
depending on the object type.
------------------------------------------------------------
"""


class Type(Enum):
    XLRD = (1,)
    OPENPYXL = 2


def wb_type(wb):
    if isinstance(wb, xlrd.book.Book):
        return Type.XLRD
    elif isinstance(wb, openpyxl.workbook.workbook.Workbook):
        return Type.OPENPYXL


def sheet_type(sheet):
    if isinstance(sheet, xlrd.sheet.Sheet):
        return Type.XLRD
    elif isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        return Type.OPENPYXL


def num_sheets(wb):
    match wb_type(wb):
        case Type.XLRD:
            return wb.nsheets
        case Type.OPENPYXL:
            return len(wb.worksheets)


def sheet_by_index(wb, idx):
    match wb_type(wb):
        case Type.XLRD:
            sheet = wb.sheet_by_index(idx)
            return sheet, sheet.name
        case Type.OPENPYXL:
            sheet = wb.worksheets[idx]
            return sheet, sheet.title


def hex_from_bgx(wb, bgx):
    match wb_type(wb):
        case Type.XLRD:
            (r, g, b) = wb.colour_map[bgx]
            hx = matplotlib.colors.to_hex((r / 255, g / 255, b / 255))
            return hx
        case Type.OPENPYXL:
            # hx = hex(bgx)
            # (r, g, b) = tuple(int(hx[i:i+2], 16) for i in (0, 2, 4))
            # hx = matplotlib.colors.to_hex((r / 255, g / 255, b / 255))
            # Colors = styles.colors.COLOR_INDEX
            # result = str(Colors[hx])
            # result = "#"+result[2:]
            # TODO: Haven't yet figured out how to convert color index
            #       to hex color. Just return 'white' for now.
            return "#FFFFFF"


def sheet_cols_rows(sheet):
    match sheet_type(sheet):
        case Type.XLRD:
            return sheet.nrows, sheet.ncols
        case Type.OPENPYXL:
            return sheet.max_row, sheet.max_column


def sheet_cell_details(sheet, wb, r, c):
    """
    Gets the sheet's cell[r][c] properties
    """
    match sheet_type(sheet):
        case Type.XLRD:
            col_type = sheet.cell_type(r, c)
            cell_obj = sheet.cell(r, c)
            xfx = sheet.cell_xf_index(r, c)
            xf = wb.xf_list[xfx]
            bgx = xf.background.pattern_colour_index
            empty_or_ws = (
                (col_type == xlrd.XL_CELL_EMPTY) or (cell_obj == "") or bgx == 64
            )
            return empty_or_ws, bgx, cell_obj.value
        case Type.OPENPYXL:
            cell = sheet.cell(r + 1, c + 1)
            bgx = int(str(cell.fill.start_color.index), base=16)
            empty_or_ws = bgx == 0
            return empty_or_ws, bgx, cell.value


"""
------------------------------------------------------------
                        Main logic
------------------------------------------------------------
"""


def parse_input_workbook(in_wb, sheet_num, break_on_white):
    """
    Parse the given sheet from the input workbook and store
    the data in the parsed dictionary.
    in_wb - input workbook
    sheet - the sheet to process
    sheet_num - the sheet number
    break_on_white - break the color into a new column when seeing
    a white space.
    """
    global input_dic_parsed, cur_column_per_color

    sheet, name = sheet_by_index(in_wb, sheet_num)
    logging.info("Processing sheet: %s" % name)
    # Get the total number of rows and columns in this sheet
    rows, cols = sheet_cols_rows(sheet)
    logging.debug("Max Rows: %d, cols: %d", rows, cols)
    for col in range(0, cols):  # Iterate through columns in this sheet
        color_seen_in_this_col = {}
        # For each column, iterate through the rows in that column
        last_non_white_bgx = -1
        white_space_encountered = False
        for row in range(0, rows):
            # Get cell details of [row, col]
            empty_or_ws, bgx, value = sheet_cell_details(sheet, in_wb, row, col)
            if empty_or_ws:
                logging.debug("[%d,%d] empty cell" % (row + 1, col + 1))
                white_space_encountered = True
                continue  # skip empty cells
            if break_on_white and white_space_encountered and last_non_white_bgx != -1:
                logging.debug("breaking on white for color: %d" % (last_non_white_bgx))
                cur_column_per_color[last_non_white_bgx] += 1
            last_non_white_bgx = bgx
            white_space_encountered = False
            # record all the colors that are seen in this column
            if bgx not in color_seen_in_this_col.keys():
                color_seen_in_this_col[bgx] = 1
            if bgx not in cur_column_per_color.keys():
                cur_column_per_color[bgx] = -1
            column = cur_column_per_color[bgx]
            column += 1
            logging.debug("[%d ,%d] bgx: %d" % (row + 1, col + 1, bgx))
            input_dic_parsed[sheet_num][bgx][column].append(value)
        # logging.debug('colors in this col %s', color_seen_in_this_col)
        # for all the colors that were there in this column, increment
        # the column count, so that the next hit of this color can
        # go to the next output column.
        for colors in color_seen_in_this_col:
            cur_column_per_color[colors] += 1
        logging.debug("current column per color %s" % (cur_column_per_color))
    logging.debug("parsed dict %s" % (input_dic_parsed))


def generate_output(out_wb, in_wb):
    global cur_column_per_color, input_dic_parsed

    # cur_column_per_color = {}
    cell_format = out_wb.add_format()
    for sheet_num in input_dic_parsed:
        for color in list(input_dic_parsed[sheet_num]):
            logging.debug("color %s, sheet_num: %d", color, sheet_num)
            if color not in cur_column_per_color.keys():
                cur_column_per_color[color] = 0
            # print 'Color code: %s, col is %s' % (color, col)
            for column in input_dic_parsed[sheet_num][color]:
                row = 0
                for value in input_dic_parsed[sheet_num][color][column]:
                    logging.debug("column %s, value: %s" % (column, value))
                    sheet_name = "Sheet" + str(color_to_sheet_num_map[color])
                    sheet = out_wb.get_worksheet_by_name(sheet_name)
                    if row == 0:
                        # first entry is the input sheet name
                        name = sheet_by_index(in_wb, sheet_num)[1]
                        sheet.write(row, column, name, cell_format)
                        row += 1
                    sheet.write(row, column, value, cell_format)
                    row += 1


def write_output_file(out_wb):
    while True:
        try:
            out_wb.close()
            break
        except xlsxwriter.exceptions.FileCreateError:
            s_continue = app.yesno(
                "Permssion Denied",
                "Output file is open in Excel and cannot be edited. Please close the file in Excel.\nRetry after closing?\nClick yes to retry or no to abort operation",
            )

            if s_continue != True:
                return False

    return True


def bg_to_hex(rgb):
    (r, g, b) = rgb
    return ("#{:X}{:X}{:X}").format(r, g, b)


def sort(input_file, output_file, break_on_ws):
    global logger, color_to_sheet_num_map, input_dic_parsed, cur_column_per_color

    input_dic_parsed.clear()
    cur_column_per_color.clear()

    logger.info(
        "Processing input file: %s, break on whitespace: %d", input_file, break_on_ws
    )

    color_to_sheet_num_map.clear()
    ext = os.path.splitext(os.path.basename(input_file))[1]
    if ext == ".xls":
        in_wb = xlrd.open_workbook(input_file, formatting_info=True)
    elif ext == ".xlsx":
        in_wb = openpyxl.load_workbook(input_file)
    else:
        logger.error("Unsupported extension (%s). Only .xls or xlsx supported.", ext)
        return False

    for sheet_num in range(0, num_sheets(in_wb)):
        parse_input_workbook(in_wb, sheet_num, break_on_ws)

    out_wb = xlsxwriter.Workbook(output_file)
    logging.info("Total distinct sheets = %s", len(input_dic_parsed))
    color_sheet_map_count = 1
    for sheet_num in input_dic_parsed:
        for color in input_dic_parsed[sheet_num]:
            if color not in color_to_sheet_num_map.keys():
                ws = out_wb.add_worksheet()
                hex = hex_from_bgx(in_wb, color)
                ws.set_tab_color(hex)
                color_to_sheet_num_map[color] = color_sheet_map_count
                color_sheet_map_count += 1

    generate_output(out_wb, in_wb)
    logging.info("Output file written: %s" % output_file)
    return write_output_file(out_wb)


"""
------------------------------------------------------------
                UI related stuff
------------------------------------------------------------
"""


def line():
    Text(
        app,
        "--------------------------------------------------------------------------",
    )


def select_input_file():
    global input_file, input_folder, output_file

    if input_folder:
        open_folder = input_folder
    else:
        open_folder = "."
    input_file = app.select_file(
        folder=open_folder, filetypes=[("Excel files", ".xlsx .xls")]
    )
    if not input_file:
        return

    # Remember the input folder for next time.
    input_folder = os.path.dirname(input_file)
    file_name_without_ext, ext = os.path.splitext(os.path.basename(input_file))
    input_file_text_box.value = os.path.basename(input_file)
    output_file = os.path.join(input_folder, "o_" + file_name_without_ext + ".xlsx")
    open_output_file_button.enabled = False


def break_on_whitespace_selection():
    global break_on_white

    if break_on_ws_checkbox.value == 1:
        break_on_white = True
    else:
        break_on_white = False


def set_open_output_file_after_sort():
    global open_output_file_once_sorted

    if open_output_file_after_sort_checkbox.value == 1:
        open_output_file_once_sorted = True
    else:
        open_output_file_once_sorted = False


def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])


def open_output_file():
    global output_file

    if not output_file:
        app.warn(
            "Uh oh!",
            "Output file does not exist yet. Please select an input file and sort to create an output file first!",
        )
        return

    open_file(output_file)


def sort_cick():
    global input_file, output_file, break_on_white

    open_output_file_button.enabled = False
    if not input_file:
        app.warn(
            "Uh oh!",
            "No input file selected to color sort. Please select an input excel file and try again.",
        )
        return

    if sort(input_file, output_file, break_on_white) == True:
        open_output_file_button.enabled = True
        if open_output_file_once_sorted:
            open_file(output_file)
    else:
        open_output_file_button.enabled = False


def print_help():
    """
    Display help
    """
    print("\nHelp/Usage:\n")
    print("python colorsort.py -v -h\n")
    print("where:")
    print("-v Enable verbose logging.")
    print("-h Display help.")


if __name__ == "__main__":
    """
    Main entry point
    """

    log_level = logging.INFO
    try:
        opts, args = getopt.getopt(sys.argv[1:], "vh:")
    except getopt.GetoptError:
        print_help()
    for opt, arg in opts:
        if opt in ("-v"):
            log_level = logging.DEBUG
        else:
            print_help()

    logging.basicConfig(
        filename=OUTPUT_LOG_FILE, level=log_level, filemode="w+", format=""
    )
    logger = logging.getLogger(__name__)

    # Main app
    app = App("", height=450, width=400)

    # App name box
    title = Text(app, text="Excel color sort", size=16, font="Arial Bold", width=30)
    title.bg = "white"
    line()

    # Select input folder button
    Text(app, "Select Excel File --> Sort", font="Verdana bold")
    line()
    input_file_button = PushButton(
        app, command=select_input_file, text="Select Excel File", width=26
    )
    input_file_button.tk.config(font=("Verdana bold", 10))
    input_file_button.bg = "#ff9933"
    input_file_button.text_color = "white"

    # Box to display the input folder
    line()
    input_file_text_box = TextBox(app)
    # Non editable
    input_file_text_box.disable()
    input_file_text_box.width = 26
    input_file_text_box.font = "Verdana bold"
    input_file_text_box.text_size = 12
    line()

    # Process input button
    process_button = PushButton(app, text="Color Sort", command=sort_cick, width=26)
    process_button.tk.config(font=("Verdana bold", 14))
    process_button.bg = "#0099cc"
    process_button.text_color = "white"
    break_on_ws_checkbox = CheckBox(
        app,
        text="Break on whitespace into next column",
        command=break_on_whitespace_selection,
    )
    break_on_ws_checkbox.value = break_on_white
    open_output_file_after_sort_checkbox = CheckBox(
        app,
        text="Automatically open output file",
        command=set_open_output_file_after_sort,
    )
    open_output_file_after_sort_checkbox.value = open_output_file_once_sorted
    line()

    # Browse output folder button
    open_output_file_button = PushButton(
        app, command=open_output_file, text="Open output file", width=20
    )
    open_output_file_button.tk.config(font=("Verdana bold", 10))
    open_output_file_button.enabled = False
    line()

    # Display the app
    app.display()
